import csv
import hashlib
import io
import json
import logging
import os
import re
import secrets
import sqlite3
import time

logging.basicConfig(level=logging.INFO)

import anthropic
from flask import Flask, Response, jsonify, redirect, render_template, request, session, url_for
from functools import wraps

from analyzer import analyze_niche
from meli_api import (
    derive_keyword, fetch_orders_total, fetch_sales_by_item, get_categories,
    get_fulfillment_stock, get_item_position, get_items_catalog_ids, get_items_prices,
    get_my_store_items, get_my_user_id, get_subcategories, sample_subcategory, search_meli,
)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))
app.config["MAX_CONTENT_LENGTH"] = 120 * 1024 * 1024  # 120 MB
# DB_PATH configurable para apuntar a un volumen persistente en Railway
DB_PATH = os.environ.get("DB_PATH") or os.path.join(os.path.dirname(os.path.abspath(__file__)), "melichnicho.db")


def init_db():
    with sqlite3.connect(DB_PATH) as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS searches (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                query       TEXT,
                category_id TEXT,
                filters     TEXT,
                results_count INTEGER,
                niche_stats TEXT,
                timestamp   DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS search_results (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                search_id        INTEGER,
                item_id          TEXT,
                title            TEXT,
                price            REAL,
                sold_quantity    INTEGER,
                seller_id        TEXT,
                seller_name      TEXT,
                opportunity_score REAL,
                free_shipping    INTEGER,
                permalink        TEXT,
                thumbnail        TEXT,
                FOREIGN KEY (search_id) REFERENCES searches(id)
            );
        """)


def init_users_table():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id       INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        admin_user = os.environ.get("ADMIN_USER")
        admin_pass = os.environ.get("ADMIN_PASSWORD")
        logging.info(f"ADMIN_USER env var present: {bool(admin_user)}, ADMIN_PASSWORD env var present: {bool(admin_pass)}")
        if admin_user and admin_pass:
            pw_hash = hashlib.sha256(admin_pass.encode()).hexdigest()
            logging.info(f"Upserting admin user: '{admin_user}'")
            conn.execute(
                "INSERT INTO users (username, password_hash) VALUES (?, ?)"
                " ON CONFLICT(username) DO UPDATE SET password_hash = excluded.password_hash",
                (admin_user, pw_hash),
            )
            count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
            logging.info(f"Users in DB: {count}")


def init_products_table():
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS product_config (
                id                   INTEGER PRIMARY KEY AUTOINCREMENT,
                item_ids             TEXT NOT NULL DEFAULT '[]',
                landed_cost_ars      REAL,
                costo_usd            REAL,
                proyeccion_mes       INTEGER,
                keyword              TEXT,
                position             INTEGER,
                position_total       INTEGER,
                position_ts          DATETIME,
                position_competitors TEXT,
                updated_at           DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cols = {row[1] for row in conn.execute("PRAGMA table_info(product_config)")}
        if "costo_usd" not in cols:
            conn.execute("ALTER TABLE product_config ADD COLUMN costo_usd REAL")


init_db()
init_users_table()
init_products_table()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            if request.path.startswith("/api/"):
                return jsonify({"error": "No autorizado"}), 401
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


@app.route("/login", methods=["GET", "POST"])
def login():
    if "user" in session:
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        pw_hash = hashlib.sha256(password.encode()).hexdigest()
        with sqlite3.connect(DB_PATH) as conn:
            row = conn.execute(
                "SELECT id FROM users WHERE username = ? AND password_hash = ?",
                (username, pw_hash),
            ).fetchone()
        logging.info(f"Login attempt: user='{username}', hash='{pw_hash[:12]}...'")
        if row:
            session["user"] = username
            logging.info(f"Login success: {username}")
            return redirect(url_for("index"))
        logging.info(f"Login failed: no matching user/password")
        error = "Usuario o contraseña incorrectos"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    return render_template("index.html")


@app.route("/nubi-results")
@login_required
def nubi_results():
    return render_template("nubi_results.html")


@app.route("/api/categories")
@login_required
def categories():
    return jsonify(get_categories())


@app.route("/api/categories/<category_id>/children")
@login_required
def category_children(category_id):
    return jsonify(get_subcategories(category_id))


# Sugerencias curadas para cuando la DB todavía no tiene historial propio
FALLBACK_SUGGESTIONS = [
    "detector fugas de gas", "microscopio usb", "controlador temperatura stc-1000",
    "sensor monoxido de carbono", "protector de tension", "masajeador cervical",
    "regulador solar", "multimetro digital",
]


@app.route("/api/suggestions")
@login_required
def suggestions():
    # Nichos populares: búsquedas previas con resultados, rankeadas por
    # frecuencia de uso y baja competencia (unique_sellers del niche_stats)
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT query, MAX(category_id) AS category_id, COUNT(*) AS veces,"
            " MAX(timestamp) AS last_ts, MAX(niche_stats) AS niche_stats"
            " FROM searches"
            " WHERE results_count > 0 AND query != '' AND query != IFNULL(category_id, '')"
            " GROUP BY LOWER(query) ORDER BY last_ts DESC LIMIT 60"
        ).fetchall()

    scored = []
    for r in rows:
        try:
            stats = json.loads(r["niche_stats"] or "{}")
        except ValueError:
            stats = {}
        sellers = stats.get("unique_sellers") or 0
        score = r["veces"] * 10 + max(0, 40 - sellers)
        scored.append({
            "query": r["query"],
            "competition_level": stats.get("competition_level", ""),
            "score": score,
        })
    scored.sort(key=lambda s: s["score"], reverse=True)
    out = scored[:8]

    for q in FALLBACK_SUGGESTIONS:
        if len(out) >= 8:
            break
        if not any(s["query"].lower() == q for s in out):
            out.append({"query": q, "competition_level": "", "score": 0})
    return jsonify(out)


@app.route("/api/search", methods=["POST"])
@login_required
def search():
    data = request.get_json() or {}
    query = (data.get("query") or "").strip()
    category_id = (data.get("category_id") or "").strip()

    if not query and not category_id:
        return jsonify({"error": "Ingresá una keyword o seleccioná una categoría."}), 400

    filters = {
        "min_price": data.get("min_price"),
        "max_price": data.get("max_price"),
        "min_sold": data.get("min_sold", 0),
        "max_sellers": data.get("max_sellers"),
        "free_shipping": data.get("free_shipping", False),
    }

    raw = search_meli(query=query, category_id=category_id)
    if "error" in raw:
        return jsonify(raw), 502

    items, niche_stats, seller_ranking = analyze_niche(raw["items"], filters)

    max_sellers = filters.get("max_sellers")
    if max_sellers and niche_stats.get("unique_sellers", 0) > int(max_sellers):
        niche_stats["competition_warning"] = (
            f"Nicho con {niche_stats['unique_sellers']} vendedores "
            f"(tu límite: {int(max_sellers)})"
        )

    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute(
            "INSERT INTO searches (query, category_id, filters, results_count, niche_stats)"
            " VALUES (?,?,?,?,?)",
            (query or category_id, category_id, json.dumps(filters),
             len(items), json.dumps(niche_stats)),
        )
        search_id = c.lastrowid
        c.executemany(
            "INSERT INTO search_results"
            " (search_id, item_id, title, price, sold_quantity, seller_id, seller_name,"
            "  opportunity_score, free_shipping, permalink, thumbnail)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            [
                (search_id, i["id"], i["title"], i["price"], i.get("sold_quantity", 0),
                 str(i.get("seller_id", "")), i.get("seller_name", ""),
                 i.get("opportunity_score", 0), 1 if i.get("free_shipping") else 0,
                 i.get("permalink", ""), i.get("thumbnail", ""))
                for i in items
            ],
        )

    return jsonify({
        "search_id": search_id,
        "items": items,
        "niche_stats": niche_stats,
        "seller_ranking": seller_ranking,
    })


@app.route("/api/history")
@login_required
def history():
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT id, query, category_id, results_count, timestamp"
            " FROM searches ORDER BY timestamp DESC LIMIT 20"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/export/<int:search_id>")
@login_required
def export_csv(search_id):
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT * FROM search_results WHERE search_id = ?"
            " ORDER BY opportunity_score DESC",
            (search_id,),
        ).fetchall()

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Título", "Precio (ARS)", "Vendidos", "Vendedor",
                "Score Oportunidad", "Envío Gratis", "Link"])
    for r in rows:
        w.writerow([
            r["title"], r["price"], r["sold_quantity"], r["seller_name"],
            r["opportunity_score"], "Sí" if r["free_shipping"] else "No", r["permalink"],
        ])

    return Response(
        "﻿" + buf.getvalue(),  # BOM for Excel
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=melichnicho_{search_id}.csv"},
    )


@app.route("/api/discover", methods=["POST"])
@login_required
def discover():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada en el servidor."}), 500

    data = request.get_json() or {}
    category_id = (data.get("category_id") or "").strip()
    if not category_id:
        return jsonify({"error": "Seleccioná una categoría."}), 400

    subcats = get_subcategories(category_id)
    if not subcats:
        return jsonify({"error": "No se encontraron subcategorías para analizar."}), 404

    subcats_with_data = []
    for sc in subcats[:8]:
        metrics = sample_subcategory(sc["id"])
        if metrics:
            subcats_with_data.append({
                "id": sc["id"],
                "name": sc["name"],
                "total_items_in_category": sc.get("total_items_in_this_category", 0),
                **metrics,
            })
        time.sleep(0.3)

    if not subcats_with_data:
        return jsonify({"error": "No se pudieron obtener datos de las subcategorías."}), 502

    datos_str = json.dumps(subcats_with_data, ensure_ascii=False, indent=2)
    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Te paso datos de subcategorías dentro de una categoría.\n"
        "Analizá y encontrá las MEJORES OPORTUNIDADES de nicho, priorizando:\n"
        "1. Pocos vendedores únicos (baja competencia)\n"
        "2. Buena demanda (ventas promedio decentes)\n"
        "3. Precios que permitan margen\n\n"
        "Devolvé un informe en español con:\n"
        "- Top 3-5 nichos recomendados, ordenados de mejor a peor oportunidad\n"
        "- Para cada uno: nombre de la subcategoría, por qué es buena oportunidad, "
        "nivel de competencia, rango de precios sugerido, y un veredicto "
        "(🟢 Entrar / 🟡 Evaluar / 🔴 Evitar)\n"
        "- Un resumen final con tu recomendación principal\n\n"
        f"Datos de subcategorías:\n{datos_str}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        analysis_text = response.content[0].text
        return jsonify({
            "analysis": analysis_text,
            "subcategories_analyzed": len(subcats_with_data),
            "data": subcats_with_data,
        })
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/analyze", methods=["POST"])
@login_required
def analyze_ai():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada en el servidor."}), 500

    data = request.get_json() or {}
    niche_data = data.get("niche_data", {})
    datos_str = json.dumps(niche_data, ensure_ascii=False, indent=2)

    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Analizá este nicho y devolvé un informe breve en español con:\n"
        "1. Resumen del mercado\n"
        "2. Nivel de competencia (bajo/medio/alto)\n"
        "3. Oportunidad de entrada\n"
        "4. Precio recomendado para entrar\n"
        "5. Veredicto final: una de estas tres opciones exactas → "
        "🟢 Recomendado / 🟡 Evaluar / 🔴 Evitar\n\n"
        f"Datos del nicho: {datos_str}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        )
        analysis_text = response.content[0].text

        tail = analysis_text[-600:].lower()
        if "🟢" in analysis_text[-600:] or "recomendado" in tail:
            verdict = "green"
        elif "🔴" in analysis_text[-600:] or "evitar" in tail:
            verdict = "red"
        else:
            verdict = "yellow"

        return jsonify({"analysis": analysis_text, "verdict": verdict})
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/rt-upload", methods=["POST"])
@login_required
def rt_upload():
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl no instalado en el servidor."}), 500

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No se recibió ningún archivo."}), 400

    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el archivo: {e}"}), 400

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "El archivo está vacío."}), 400

    # Detectar fila de encabezado buscando columnas clave
    header_row = None
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        row_lower = [str(c).lower().strip() if c is not None else "" for c in row]
        if any("unidades" in c for c in row_lower) or any("publicacion" in c for c in row_lower):
            header_row = row_lower
            header_idx = i
            break

    if header_row is None:
        return jsonify({"error": "No se encontraron las columnas esperadas (Unidades vendidas, Publicaciones)."}), 400

    def col_idx(keywords):
        for i, h in enumerate(header_row):
            if any(k in h for k in keywords):
                return i
        return None

    idx_title   = col_idx(["publicacion", "título", "titulo", "producto", "nombre"])
    idx_seller  = col_idx(["vendedor", "seller"])
    idx_price   = col_idx(["precio"])
    idx_units   = col_idx(["unidades"])
    idx_revenue = col_idx(["facturacion", "facturación", "revenue"])

    items = []
    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue
        def val(i):
            if i is None or i >= len(row):
                return None
            v = row[i]
            return v

        title = val(idx_title)
        if not title:
            continue

        # Limpiar número: "1.500" → 1500, "+1.500" → 1500
        def to_num(v):
            if v is None:
                return 0
            s = re.sub(r"[^\d]", "", str(v))
            return int(s) if s else 0

        def to_float(v):
            if v is None:
                return 0.0
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).replace("$", "").replace("+", "").strip()
            s = s.replace(".", "").replace(",", ".")
            try:
                return float(s)
            except:
                return 0.0

        items.append({
            "title":    str(title).strip(),
            "seller":   str(val(idx_seller)).strip() if val(idx_seller) else "",
            "price":    to_float(val(idx_price)),
            "units":    to_num(val(idx_units)),
            "revenue":  to_float(val(idx_revenue)),
        })

    if not items:
        return jsonify({"error": "No se encontraron filas de datos en el archivo."}), 400

    items.sort(key=lambda x: x["units"], reverse=True)
    return jsonify({"items": items, "total": len(items)})


@app.route("/api/rt-analyze", methods=["POST"])
@login_required
def rt_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada en el servidor."}), 500

    data = request.get_json() or {}
    items = data.get("items", [])
    if not items:
        return jsonify({"error": "No hay datos para analizar."}), 400

    items_str = json.dumps(items[:30], ensure_ascii=False, indent=2)

    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Te paso el ranking de publicaciones de una categoría exportado desde Real Trends.\n"
        "Cada item tiene: título, vendedor, precio promedio, unidades vendidas, facturación.\n\n"
        "Analizá estos datos y detectá OPORTUNIDADES DE NEGOCIO para un vendedor que quiere entrar o crecer en esta categoría.\n\n"
        "Tu análisis debe incluir:\n"
        "1. **Resumen del mercado** — tamaño, concentración, quiénes dominan\n"
        "2. **Segmentos de precio** — identificá rangos de precio y cuáles tienen más demanda\n"
        "3. **Oportunidades detectadas** — productos o nichos dentro de la categoría con:\n"
        "   - Alta demanda (unidades vendidas) pero pocos competidores dominantes\n"
        "   - Segmentos de precio sin jugadores fuertes\n"
        "   - Variantes de producto sub-representadas en el top\n"
        "4. **Estrategia recomendada** — precio de entrada, tipo de producto, diferenciación\n"
        "5. **Veredicto** — 🟢 Oportunidad clara / 🟡 Evaluar / 🔴 Mercado saturado\n\n"
        f"Datos del ranking:\n{items_str}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        return jsonify({"analysis": response.content[0].text})
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/nubi-upload", methods=["POST"])
@login_required
def nubi_upload():
    import statistics
    from collections import defaultdict

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No se recibió ningún archivo."}), 400

    try:
        stream = io.StringIO(f.stream.read().decode("utf-8", errors="replace"))
        reader = csv.DictReader(stream)
        rows = list(reader)
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el CSV: {e}"}), 400

    if not rows:
        return jsonify({"error": "El archivo está vacío."}), 400

    def fnum(v):
        try: return float(v or 0)
        except: return 0.0

    def fint(v):
        try: return int(float(v or 0))
        except: return 0

    # ─── Aggregación por subcategoría ─────────────────────
    subcats = defaultdict(lambda: {
        "units": 0, "revenue": 0, "listings": 0,
        "prices": [], "full": 0, "fship": 0,
        "sellers": defaultdict(int),
        "products": defaultdict(lambda: {"units": 0, "price": 0, "seller": ""}),
    })

    price_buckets = {"<10k": [0,0], "10k-25k": [0,0], "25k-50k": [0,0],
                     "50k-100k": [0,0], "100k-200k": [0,0], "+200k": [0,0]}

    total_units = 0
    total_revenue = 0.0
    unique_sellers_global = set()

    for r in rows:
        cat = (r.get("Categoria_Nivel_4") or "").strip()
        if not cat or cat == "-":
            cat = (r.get("Categoria_Nivel_3") or "").strip() or "Otros"

        u    = fint(r.get("Unidades_Vendidas"))
        rev  = fnum(r.get("Monto_Vendido_Moneda_Local"))
        price = fnum(r.get("PrecioMonedaLocal"))
        seller = r.get("Nickname_Vendedor", "")
        title  = (r.get("Titulo_Publicacion") or "")[:80].strip()

        total_units   += u
        total_revenue += rev
        unique_sellers_global.add(seller)

        d = subcats[cat]
        d["units"]    += u
        d["revenue"]  += rev
        d["listings"] += 1
        if price > 0: d["prices"].append(price)
        if r.get("OfreceFull") == "Si":           d["full"]  += 1
        if r.get("Ofrece_Envio_Gratis") == "true": d["fship"] += 1
        d["sellers"][seller] += u
        p = d["products"][title]
        p["units"] += u
        if price > p["price"]: p["price"] = price
        p["seller"] = seller

        # price buckets
        if price < 10_000:      price_buckets["<10k"][0] += u;    price_buckets["<10k"][1] += 1
        elif price < 25_000:    price_buckets["10k-25k"][0] += u; price_buckets["10k-25k"][1] += 1
        elif price < 50_000:    price_buckets["25k-50k"][0] += u; price_buckets["25k-50k"][1] += 1
        elif price < 100_000:   price_buckets["50k-100k"][0] += u; price_buckets["50k-100k"][1] += 1
        elif price < 200_000:   price_buckets["100k-200k"][0] += u; price_buckets["100k-200k"][1] += 1
        else:                   price_buckets["+200k"][0] += u;    price_buckets["+200k"][1] += 1

    # ─── Serializar resultado ──────────────────────────────
    subcat_list = []
    for name, d in subcats.items():
        n = d["listings"]
        u = d["units"]
        top_sellers = sorted(d["sellers"].items(), key=lambda x: -x[1])[:3]
        top3_units  = sum(x[1] for x in top_sellers)
        top_products = sorted(d["products"].items(), key=lambda x: -x[1]["units"])[:5]

        subcat_list.append({
            "name":            name,
            "listings":        n,
            "unique_sellers":  len(d["sellers"]),
            "total_units":     u,
            "total_revenue":   round(d["revenue"]),
            "avg_price":       round(statistics.mean(d["prices"])) if d["prices"] else 0,
            "median_price":    round(statistics.median(d["prices"])) if d["prices"] else 0,
            "pct_full":        round(d["full"] / n * 100) if n else 0,
            "pct_free_ship":   round(d["fship"] / n * 100) if n else 0,
            "top3_concentration": round(top3_units / u * 100) if u else 0,
            "top_sellers":     [{"seller": s, "units": v} for s, v in top_sellers],
            "top_products":    [{"title": t, "units": p["units"], "price": round(p["price"]), "seller": p["seller"]} for t, p in top_products],
        })

    subcat_list.sort(key=lambda x: -x["total_units"])

    # top 15 productos globales
    all_products = defaultdict(lambda: {"units": 0, "price": 0, "seller": "", "cat": ""})
    for r in rows:
        title = (r.get("Titulo_Publicacion") or "")[:80].strip()
        u = fint(r.get("Unidades_Vendidas"))
        price = fnum(r.get("PrecioMonedaLocal"))
        all_products[title]["units"] += u
        if price > all_products[title]["price"]: all_products[title]["price"] = price
        all_products[title]["seller"] = r.get("Nickname_Vendedor", "")
        cat = (r.get("Categoria_Nivel_4") or r.get("Categoria_Nivel_3") or "").strip()
        all_products[title]["cat"] = cat

    top_global = sorted(all_products.items(), key=lambda x: -x[1]["units"])[:15]

    meta = {
        "category_name": (rows[0].get("Categoria_Nivel_2") or rows[0].get("Categoria_Nivel_1") or "").strip(),
        "period": (rows[0].get("Mes") or "")[:7],
        "total_listings": len(rows),
        "total_units": total_units,
        "total_revenue_ars": round(total_revenue),
        "unique_sellers": len(unique_sellers_global),
    }

    return jsonify({
        "meta": meta,
        "subcategories": subcat_list,
        "top_products": [{"title": t, **p} for t, p in top_global],
        "price_segments": {k: {"units": v[0], "listings": v[1]} for k, v in price_buckets.items()},
    })


@app.route("/api/nubi-analyze", methods=["POST"])
@login_required
def nubi_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada en el servidor."}), 500

    data = request.get_json() or {}
    agg  = data.get("data", {})
    if not agg:
        return jsonify({"error": "Sin datos para analizar."}), 400

    meta     = agg.get("meta", {})
    subcats  = agg.get("subcategories", [])[:20]
    top_prods = agg.get("top_products", [])
    segments = agg.get("price_segments", {})

    summary = json.dumps({
        "meta": meta,
        "subcategories": subcats,
        "top_15_productos": top_prods,
        "segmentos_de_precio": segments,
    }, ensure_ascii=False, indent=1)

    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Te paso datos agregados de Nubimetrics para una categoría completa.\n\n"
        "IMPORTANTE — formato de números:\n"
        "- Expresá todos los montos en formato abreviado: M para millones, B para miles de millones.\n"
        "  Ejemplos correctos: $5.9B, $520M, $14.3M, $103k. NUNCA escribas números largos como $5.878.334.409.\n"
        "- Las unidades vendidas también abrevialas si superan mil: 14.4k, 120k.\n\n"
        "Los datos incluyen por subcategoría:\n"
        "- listings, unique_sellers, total_units, total_revenue, avg_price, median_price\n"
        "- pct_full (% publicaciones con MeLi Full), pct_free_ship, top3_concentration (% de unidades de los 3 top vendedores)\n"
        "- top_products y top_sellers\n\n"
        "Analizá estos datos y producí un informe de oportunidades de negocio con:\n\n"
        "## 1. Resumen del mercado\n"
        "Tamaño total, categorías más grandes, concentración general.\n\n"
        "## 2. Top 5 Oportunidades de Nicho\n"
        "Para cada oportunidad usá este formato exacto:\n"
        "### [Nombre de la subcategoría]\n"
        "- Por qué es oportunidad\n"
        "- Competencia: vendedores, concentración top 3\n"
        "- Precio recomendado para entrar\n"
        "- Estrategia: Full / envío gratis\n"
        "- Veredicto: 🟢 Entrar / 🟡 Evaluar / 🔴 Evitar\n\n"
        "## 3. Segmentos de precio con más demanda\n"
        "Qué rango de precio concentra más unidades y por qué.\n\n"
        "## 4. Advertencias\n"
        "Subcategorías saturadas o con barreras de entrada altas.\n\n"
        "## 5. Recomendación final\n"
        "La mejor oportunidad concreta para un vendedor nuevo o en crecimiento.\n\n"
        f"Datos:\n{summary}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}],
        )
        return jsonify({"analysis": response.content[0].text})
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/nubi-export", methods=["POST"])
@login_required
def nubi_export():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl no instalado."}), 500

    body = request.get_json() or {}
    data     = body.get("data", {})
    analysis = body.get("analysis", "")

    if not data:
        return jsonify({"error": "Sin datos para exportar."}), 400

    meta      = data.get("meta", {})
    subcats   = data.get("subcategories", [])
    top_prods = data.get("top_products", [])
    segments  = data.get("price_segments", {})

    wb = openpyxl.Workbook()

    # ─── Estilos ──────────────────────────────────────────
    hdr_font    = Font(bold=True, color="000000")
    hdr_fill    = PatternFill("solid", fgColor="FFE600")
    title_font  = Font(bold=True, size=13, color="FFE600")
    label_font  = Font(bold=True, color="E0E0E0")
    dark_fill   = PatternFill("solid", fgColor="16213E")
    green_fill  = PatternFill("solid", fgColor="1B5E20")
    orange_fill = PatternFill("solid", fgColor="E65100")
    red_fill    = PatternFill("solid", fgColor="B71C1C")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin        = Border(
        left=Side(style="thin", color="0F3460"),
        right=Side(style="thin", color="0F3460"),
        top=Side(style="thin", color="0F3460"),
        bottom=Side(style="thin", color="0F3460"),
    )

    def style_header_row(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = center
            cell.border = thin

    def style_data_cell(cell, fill=None):
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = thin
        if fill:
            cell.fill = fill

    # ─── Hoja 1: Resumen ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 25

    ws1["A1"] = f"📈 Nubimetrics — {meta.get('category_name', '')} | {meta.get('period', '')}"
    ws1["A1"].font = title_font

    meta_rows = [
        ("Total de listings",    f"{meta.get('total_listings', 0):,}"),
        ("Unidades vendidas",    f"{meta.get('total_units', 0):,}"),
        ("Facturación total ARS",f"${meta.get('total_revenue_ars', 0):,.0f}"),
        ("Vendedores únicos",    f"{meta.get('unique_sellers', 0):,}"),
    ]
    for i, (label, value) in enumerate(meta_rows, start=3):
        ws1.cell(row=i, column=1, value=label).font = label_font
        ws1.cell(row=i, column=2, value=value)

    ws1["A8"] = "Segmentos de precio"
    ws1["A8"].font = Font(bold=True, color="FFE600")
    ws1["A9"]  = "Rango";        ws1["B9"]  = "Unidades";  ws1["C9"] = "Listings"
    ws1.column_dimensions["C"].width = 15
    style_header_row(ws1, 9, 3)
    for i, (k, v) in enumerate(segments.items(), start=10):
        ws1.cell(row=i, column=1, value=k)
        ws1.cell(row=i, column=2, value=v.get("units", 0))
        ws1.cell(row=i, column=3, value=v.get("listings", 0))
        for c in range(1, 4):
            style_data_cell(ws1.cell(row=i, column=c))

    # ─── Hoja 2: Subcategorías ────────────────────────────
    ws2 = wb.create_sheet("Subcategorías")
    ws2.sheet_view.showGridLines = False
    headers2 = ["Subcategoría","Listings","Vendedores únicos","Unidades vendidas",
                 "Revenue ARS","Precio promedio","Precio mediano",
                 "% Full","% Envío gratis","Concentración Top3 %"]
    col_widths2 = [28,10,18,18,18,16,16,10,14,20]
    for i, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.cell(row=1, column=i, value=h)
    style_header_row(ws2, 1, len(headers2))

    for r, s in enumerate(subcats, start=2):
        conc = s.get("top3_concentration", 0)
        conc_fill = green_fill if conc <= 30 else (orange_fill if conc <= 50 else red_fill)
        values = [
            s.get("name",""), s.get("listings",0), s.get("unique_sellers",0),
            s.get("total_units",0), s.get("total_revenue",0),
            s.get("avg_price",0), s.get("median_price",0),
            s.get("pct_full",0), s.get("pct_free_ship",0), conc,
        ]
        for c, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=c, value=v)
            style_data_cell(cell, fill=conc_fill if c == 10 else None)
            if c == 10:
                cell.font = Font(bold=True, color="FFFFFF")

    # ─── Hoja 3: Top Productos ────────────────────────────
    ws3 = wb.create_sheet("Top Productos")
    ws3.sheet_view.showGridLines = False
    headers3 = ["#","Título","Vendedor","Subcategoría","Unidades vendidas","Precio ARS"]
    col_widths3 = [5, 50, 25, 25, 18, 14]
    for i, (h, w) in enumerate(zip(headers3, col_widths3), start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.cell(row=1, column=i, value=h)
    style_header_row(ws3, 1, len(headers3))

    for r, p in enumerate(top_prods, start=2):
        values = [r-1, p.get("title",""), p.get("seller",""), p.get("cat",""),
                  p.get("units",0), p.get("price",0)]
        for c, v in enumerate(values, start=1):
            style_data_cell(ws3.cell(row=r, column=c, value=v))

    # ─── Hoja 4: Análisis IA ──────────────────────────────
    if analysis:
        ws4 = wb.create_sheet("Análisis IA")
        ws4.sheet_view.showGridLines = False
        ws4.column_dimensions["A"].width = 120
        # Strip markdown and write as plain text blocks
        lines = analysis.replace("**", "").replace("##", "").replace("#", "").split("\n")
        for r, line in enumerate(lines, start=1):
            cell = ws4.cell(row=r, column=1, value=line.strip())
            cell.alignment = Alignment(wrap_text=True)
            if line.startswith("##") or (line.strip() and r <= 3):
                cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"nubimetrics_{meta.get('category_name','categoria').replace(' ','_')}_{meta.get('period','')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/sales-summary")
@login_required
def sales_summary():
    from datetime import datetime, timezone, timedelta
    tz_arg = timezone(timedelta(hours=-3))
    now = datetime.now(tz_arg)

    user_id = get_my_user_id()
    if not user_id:
        return jsonify({"error": "No se pudo obtener el usuario de MeLi."}), 502

    fmt = '%Y-%m-%dT%H:%M:%S.000-0300'
    today_from  = now.replace(hour=0, minute=0, second=0, microsecond=0).strftime(fmt)
    month_from  = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).strftime(fmt)
    now_str     = now.strftime(fmt)

    today_amt, today_cnt = fetch_orders_total(user_id, today_from, now_str)
    month_amt, month_cnt = fetch_orders_total(user_id, month_from, now_str)

    return jsonify({
        "today": {"amount": today_amt, "orders": today_cnt},
        "month": {"amount": month_amt, "orders": month_cnt},
        "as_of": now.strftime('%H:%M'),
    })


@app.route("/api/my-store")
@login_required
def my_store():
    items, error = get_my_store_items()
    if error:
        return jsonify({"error": error}), 502

    from collections import defaultdict
    items = [i for i in items if i["available_quantity"] > 0]
    active = [i for i in items if i["status"] == "active"]
    total_revenue = sum(i["revenue"] for i in items)
    cat_rev = defaultdict(int)
    for item in items:
        cat_rev[item["category_id"]] += item["revenue"]
    top_cats = sorted(cat_rev.items(), key=lambda x: -x[1])[:5]

    return jsonify({
        "items": items,
        "total": len(items),
        "active_count": len(active),
        "total_revenue_est": total_revenue,
        "top_categories": [{"cat": c, "revenue": r} for c, r in top_cats],
    })


@app.route("/api/my-store-analyze", methods=["POST"])
@login_required
def my_store_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    data = request.get_json() or {}
    items = data.get("items", [])
    target_revenue = data.get("target_revenue", 30_000_000)
    if not items:
        return jsonify({"error": "Sin datos de tienda para analizar."}), 400

    top_items = items[:50]
    total_rev = sum(i["revenue"] for i in items)
    active_count = sum(1 for i in items if i["status"] == "active")

    summary = {
        "total_publicaciones": len(items),
        "publicaciones_activas": active_count,
        "revenue_historico_total_ARS": total_rev,
        "objetivo_incremento_mensual_ARS": target_revenue,
        "top_50_por_revenue_historico": [
            {
                "titulo": i["title"],
                "precio_ARS": i["price"],
                "unidades_vendidas_historico": i["sold_quantity"],
                "revenue_estimado_ARS": i["revenue"],
                "estado": i["status"],
                "categoria_id": i["category_id"],
            }
            for i in top_items
        ],
    }

    prompt = (
        "Sos un experto en e-commerce en MercadoLibre Argentina.\n"
        "Te paso el portfolio completo de un vendedor activo en MeLi Argentina.\n\n"
        f"El vendedor tiene {len(items)} publicaciones ({active_count} activas) "
        f"con un revenue histórico estimado de ${total_rev:,.0f} ARS.\n"
        f"Su objetivo concreto es sumar ${target_revenue:,.0f} ARS MÁS por mes.\n\n"
        "Nota: 'revenue_estimado' = precio × unidades_vendidas_historico (total histórico, no mensual).\n"
        "Usá los precios y productos como referencia de categoría y ticket promedio.\n\n"
        "Analizá su portfolio y recomendá 1 a 3 productos NUEVOS que debería agregar "
        "para alcanzar ese objetivo de crecimiento mensual.\n\n"
        "Para cada producto recomendado:\n"
        "- **Producto**: nombre/descripción específica y variante\n"
        "- **Por qué encaja**: relación con lo que ya vende y demanda del mercado\n"
        "- **Precio de venta sugerido** (ARS)\n"
        "- **Unidades mensuales estimadas** para alcanzar el objetivo\n"
        "- **Revenue mensual potencial** de ese producto\n"
        "- **Veredicto**: 🟢 Alta probabilidad / 🟡 Evaluar / 🔴 Riesgo alto\n\n"
        "Al final, un **Resumen ejecutivo**:\n"
        f"- Revenue mensual potencial total de los productos recomendados vs objetivo ${target_revenue:,.0f} ARS\n"
        "- Qué producto arrancar primero y por qué\n\n"
        f"Portfolio:\n{json.dumps(summary, ensure_ascii=False, indent=1)}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=3000,
            messages=[{"role": "user", "content": prompt}],
        )
        return jsonify({"analysis": response.content[0].text})
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


@app.route("/sourcing-report")
@login_required
def sourcing_report():
    return render_template("sourcing_report.html")


@app.route("/api/sourcing-analyze", methods=["POST"])
@login_required
def sourcing_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    data = request.get_json() or {}
    products = data.get("products", [])
    target_revenue = data.get("target_revenue", 0)
    min_products = data.get("min_products", 1)
    max_products = data.get("max_products", 3)
    shipping = data.get("shipping", "courier")
    tc = data.get("tc", 1500)

    if not products:
        return jsonify({"error": "Sin datos de productos para analizar."}), 400

    shipping_label = "Courier (aéreo)" if shipping == "courier" else "Marítimo (contenedor)"
    shipping_criteria = (
        "COURIER (aéreo): priorizar productos livianos (<1kg), compactos, alto valor/peso. "
        "Evitar productos voluminosos o pesados. Ideal para electrónica pequeña, accesorios, sensores, módulos."
        if shipping == "courier"
        else
        "MARÍTIMO: pueden ser productos más grandes o pesados (hasta 20-30kg, voluminosos). "
        "Mayor plazo de entrega (30-45 días). Ideal para herramientas, equipos, muebles, productos de mayor volumen."
    )

    products_str = json.dumps(products[:50], ensure_ascii=False, indent=1)

    prompt = (
        f"Sos un experto en sourcing y e-commerce en MercadoLibre Argentina.\n"
        f"Te paso datos reales de demanda de Nubimetrics (ventas históricas de MeLi Argentina).\n\n"
        f"CRITERIOS DEL VENDEDOR:\n"
        f"- Objetivo de facturación mensual adicional: ${target_revenue:,.0f} ARS\n"
        f"- Cantidad de productos a lanzar: entre {min_products} y {max_products}\n"
        f"- Método de importación: {shipping_label}\n"
        f"  → {shipping_criteria}\n"
        f"- Tipo de cambio referencia: ${tc:,.0f} ARS/USD\n\n"
        f"DATOS DE MERCADO (productos agrupados por título, ordenados por demanda total):\n"
        f"Cada producto incluye: título, precio_promedio_ARS, total_unidades, total_revenue_ARS, "
        f"vendedores_únicos, pct_full, categoría, archivo_fuente.\n\n"
        f"{products_str}\n\n"
        f"TAREA:\n"
        f"Seleccioná los mejores {min_products} a {max_products} productos de esta lista para "
        f"que el vendedor los importe y venda en MeLi, cumpliendo:\n"
        f"1. Alcanzar ${target_revenue:,.0f} ARS/mes adicionales en total\n"
        f"2. Apto para importar vía {shipping_label}\n"
        f"3. Competencia manejable (no dominada por pocos vendedores con alta concentración)\n"
        f"4. Demanda probada en datos reales de Nubimetrics\n\n"
        f"Para cada producto recomendado usá este formato:\n"
        f"### [Nombre del producto]\n"
        f"- **Demanda del mercado**: unidades totales vendidas, revenue total, vendedores compitiendo\n"
        f"- **Mi captura estimada**: si capturo X% del mercado = Y unidades/mes = $Z ARS/mes\n"
        f"- **Precio de venta sugerido**: $X ARS\n"
        f"- **FOB estimado**: USD X–Y por unidad (China)\n"
        f"- **Apto para {shipping_label}**: peso/tamaño estimado y por qué aplica\n"
        f"- **Veredicto**: 🟢 Alta oportunidad / 🟡 Evaluar / 🔴 Evitar\n\n"
        f"Al final, una tabla resumen:\n"
        f"| Producto | Precio sugerido | Captura estimada/mes |\n"
        f"| --- | --- | --- |\n"
        f"Con una línea final indicando si el objetivo de ${target_revenue:,.0f} ARS es alcanzable."
    )

    def generate():
        try:
            client = anthropic.Anthropic(api_key=api_key)
            chunks = []
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=3000,
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                for text in stream.text_stream:
                    chunks.append(text)
                    yield " "  # keep-alive: evita timeout del proxy de Railway
            yield json.dumps({"analysis": "".join(chunks)})
        except anthropic.APIError as e:
            yield json.dumps({"error": str(e)})

    return Response(
        generate(),
        mimetype="application/json",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


@app.route("/nicho-report")
@login_required
def nicho_report():
    return render_template("nicho_report.html")


@app.route("/api/nicho-analyze", methods=["POST"])
@login_required
def nicho_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    data = request.get_json() or {}
    products = data.get("products", [])
    min_units = data.get("min_units", 50)
    max_sellers = data.get("max_sellers", 3)

    if not products:
        return jsonify({"error": "Sin datos de productos para analizar."}), 400

    products_str = json.dumps(products[:120], ensure_ascii=False, indent=1)

    prompt = (
        f"Sos un experto en e-commerce y análisis de nichos en MercadoLibre Argentina.\n"
        f"Te paso ventas reales de Nubimetrics agrupadas por título de publicación, "
        f"ordenadas por unidades vendidas en el mes.\n"
        f"Cada entrada tiene: título, categoría, unidades_mes, precio_real (monto vendido ÷ unidades, "
        f"NO el precio de lista), revenue_mes y vendedores (únicos por título exacto).\n\n"
        f"{products_str}\n\n"
        f"OBJETIVO: detectar NICHOS = productos con ALTA ROTACIÓN (≥ {min_units} unidades/mes) "
        f"y POCA COMPETENCIA (≤ {max_sellers} vendedores REALES).\n\n"
        f"IMPORTANTE — competencia real: en MeLi el mismo producto aparece con títulos distintos "
        f"según el vendedor. Antes de evaluar competencia, AGRUPÁ los títulos que sean claramente "
        f"el mismo producto (mismo tipo, marca/modelo o especificación equivalente) y sumá sus "
        f"unidades y vendedores. Un título con 1 vendedor NO es nicho si hay 10 títulos casi "
        f"iguales de otros vendedores.\n\n"
        f"TAREA: rankeá los mejores 5 a 8 nichos que cumplan los criterios DESPUÉS de agrupar. "
        f"Para cada uno usá este formato:\n"
        f"### [Nombre del nicho/producto]\n"
        f"- **Rotación total**: X unidades/mes sumando las publicaciones agrupadas\n"
        f"- **Competencia real**: X vendedores (títulos agrupados: cuáles)\n"
        f"- **Precio real de venta**: rango en ARS\n"
        f"- **Revenue del nicho**: $X ARS/mes\n"
        f"- **Por qué es nicho**: qué barrera o particularidad mantiene afuera a la competencia\n"
        f"- **Veredicto**: 🟢 Nicho claro / 🟡 Evaluar / 🔴 Descartar (con un motivo)\n\n"
        f"Al final:\n"
        f"1. Una tabla resumen: | Nicho | Unid/mes | Vendedores | Precio real | Veredicto |\n"
        f"2. Una sección '⚠️ Falsos nichos' con títulos que parecen nicho por vendedor único "
        f"pero que al agrupar tienen mucha competencia.\n"
        f"Usá formatos abreviados ($5.9M, 1.2k) y sé concreto."
    )

    def generate():
        try:
            client = anthropic.Anthropic(api_key=api_key)
            chunks = []
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=4000,
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                for text in stream.text_stream:
                    chunks.append(text)
                    yield " "  # keep-alive: evita timeout del proxy de Railway
            yield json.dumps({"analysis": "".join(chunks)})
        except anthropic.APIError as e:
            yield json.dumps({"error": str(e)})

    return Response(
        generate(),
        mimetype="application/json",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


@app.route("/comp-report")
@login_required
def comp_report():
    return render_template("comp_report.html")


@app.route("/api/comp-upload", methods=["POST"])
@login_required
def comp_upload():
    """Parsea el catálogo de UN vendedor exportado desde Nubimetrics (XLSX)."""
    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl no instalado en el servidor."}), 500

    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No se recibió ningún archivo."}), 400

    seller = os.path.splitext(f.filename or "vendedor")[0]

    try:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"error": f"No se pudo leer el archivo: {e}"}), 400

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "El archivo está vacío."}), 400

    header = [str(c).lower().strip() if c is not None else "" for c in rows[0]]

    def col_idx(keywords):
        for i, h in enumerate(header):
            if any(k in h for k in keywords):
                return i
        return None

    idx_title = col_idx(["título", "titulo"])
    idx_brand = col_idx(["marca"])
    idx_rev   = col_idx(["ventas en $", "ventas en$"])
    idx_units = col_idx(["ventas en unid", "unid"])
    idx_price = col_idx(["precio promedio", "precio"])
    idx_full  = col_idx(["fulfillment"])
    idx_cat   = col_idx(["catálogo", "catalogo"])

    if idx_title is None or idx_units is None:
        return jsonify({"error": "No parece un export de catálogo de vendedor de Nubimetrics (faltan columnas Título / Ventas en Unid.)."}), 400

    def fnum(v):
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = re.sub(r"[^\d.,]", "", str(v)).replace(".", "").replace(",", ".")
        try:
            return float(s) if s else 0.0
        except ValueError:
            return 0.0

    products = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        def val(i):
            return row[i] if i is not None and i < len(row) else None
        title = str(val(idx_title) or "").strip()
        if not title:
            continue
        units = fnum(val(idx_units))
        rev   = fnum(val(idx_rev))
        # Precio real = ventas $ / unidades; fallback al precio promedio del export
        price = round(rev / units) if units > 0 else round(fnum(val(idx_price)))
        products.append({
            "titulo":   title[:90],
            "marca":    str(val(idx_brand) or "").strip(),
            "precio":   price,
            "unidades": int(units),
            "revenue":  round(rev),
            "full":     str(val(idx_full) or "").strip().lower() == "si",
            "catalogo": str(val(idx_cat) or "").strip().lower() == "si",
        })

    if not products:
        return jsonify({"error": "No se encontraron productos en el archivo."}), 400

    products.sort(key=lambda p: p["revenue"], reverse=True)
    total_rev   = sum(p["revenue"] for p in products)
    total_units = sum(p["unidades"] for p in products)
    top10_rev   = sum(p["revenue"] for p in products[:10])

    brands = {}
    for p in products:
        b = p["marca"] or "(sin marca)"
        brands[b] = brands.get(b, 0) + p["revenue"]
    top_brands = sorted(brands.items(), key=lambda kv: kv[1], reverse=True)[:6]

    stats = {
        "revenue_mes":    total_rev,
        "unidades_mes":   total_units,
        "publicaciones":  len(products),
        "ticket":         round(total_rev / total_units) if total_units else 0,
        "top10_share":    round(top10_rev / total_rev * 100) if total_rev else 0,
        "pct_full":       round(sum(1 for p in products if p["full"]) / len(products) * 100),
        "pct_catalogo":   round(sum(1 for p in products if p["catalogo"]) / len(products) * 100),
        "marcas":         [{"marca": b, "share": round(r / total_rev * 100) if total_rev else 0} for b, r in top_brands],
    }

    return jsonify({"seller": seller, "stats": stats, "top_products": products[:20]})


@app.route("/api/comp-analyze", methods=["POST"])
@login_required
def comp_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    data = request.get_json() or {}
    sellers = data.get("sellers", [])
    tc = data.get("tc", 1500)

    if not sellers:
        return jsonify({"error": "Sin competidores para analizar."}), 400

    sellers_str = json.dumps(sellers[:5], ensure_ascii=False, indent=1)

    prompt = (
        f"Sos un experto en e-commerce en MercadoLibre Argentina y sourcing desde China.\n"
        f"Te paso el catálogo mensual (export de Nubimetrics) de uno o más COMPETIDORES. "
        f"Por cada uno: stats (revenue_mes, unidades, publicaciones, ticket, top10_share = % del "
        f"revenue concentrado en el top 10, pct_full, marcas con share) y top 20 productos con "
        f"precio real (ventas ÷ unidades — las unidades vienen redondeadas en bandas, son aproximadas).\n\n"
        f"{sellers_str}\n\n"
        f"CONTEXTO DEL ANALISTA (quien te consulta):\n"
        f"- Ya importa por courier aéreo desde China (proveedores propios de fábrica directa, "
        f"FOB reales ~1/3 del listado de Alibaba) y vende en MeLi: sensores de CO y protectores de tensión\n"
        f"- Busca productos LIVIANOS (<1kg) para courier; los voluminosos/pesados no le sirven\n"
        f"- REGLA COURIER 4×: el precio de venta ARS debe ser ≥ ~4× el FOB USD en miles para dar margen. "
        f"Cálculo: neto = precio × 0,85 (comisión MeLi) − $7.000 (envío); "
        f"landed = FOB × 1,975 × TC. Tipo de cambio actual: ${tc:,.0f} ARS/USD\n\n"
        f"TAREA — por cada competidor:\n"
        f"## [Nickname]\n"
        f"1. **Perfil**: ¿concentrado (top10_share ≥ 60%) o cola larga? ¿Marca propia o revendedor? "
        f"¿Qué estrategia usa (Full, catálogo, escalera de precios)? ¿Es un modelo a estudiar o descartar?\n"
        f"2. **Huecos atacables por courier**: productos de su catálogo donde entrar conviene. Por cada uno:\n"
        f"   - Precio real actual, unidades/mes, FOB estimado (fábrica directa, no listado Alibaba)\n"
        f"   - Verificación regla 4× con margen estimado\n"
        f"   - Veredicto: 🟢 Atacar / 🟡 Evaluar / 🔴 Evitar (con motivo: certificaciones, peso, margen)\n"
        f"3. **Jugada premium**: ¿hay oportunidad de 'escalera' (versión premium/smart de sus commodities)?\n\n"
        f"Al final una tabla resumen de TODOS los huecos:\n"
        f"| Producto | Competidor | Precio real | Unid/mes | FOB est. | Margen est. | Veredicto |\n"
        f"Y una recomendación final priorizada considerando su operación actual (CO, protectores).\n"
        f"Usá formatos abreviados ($5.9M, 1.2k) y sé concreto."
    )

    def generate():
        try:
            client = anthropic.Anthropic(api_key=api_key)
            chunks = []
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=4000,
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                for text in stream.text_stream:
                    chunks.append(text)
                    yield " "  # keep-alive: evita timeout del proxy de Railway
            yield json.dumps({"analysis": "".join(chunks)})
        except anthropic.APIError as e:
            yield json.dumps({"error": str(e)})

    return Response(
        generate(),
        mimetype="application/json",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


@app.route("/buscomp-report")
@login_required
def buscomp_report():
    return render_template("buscomp_report.html")


@app.route("/api/buscomp-analyze", methods=["POST"])
@login_required
def buscomp_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    data = request.get_json() or {}
    sellers = data.get("sellers", [])
    min_conc = data.get("min_conc", 60)
    min_units = data.get("min_units", 100)
    tc = data.get("tc", 1500)

    if not sellers:
        return jsonify({"error": "Ningún vendedor cumple los criterios — relajalos y probá de nuevo."}), 400

    sellers_str = json.dumps(sellers[:30], ensure_ascii=False, indent=1)

    prompt = (
        f"Sos un experto en e-commerce en MercadoLibre Argentina y sourcing desde China.\n"
        f"Te paso vendedores de MeLi (datos reales de Nubimetrics, un mes) que ya pasaron un filtro: "
        f"concentración top3 ≥ {min_conc}% de su revenue y producto estrella con ≥ {min_units} u/mes. "
        f"Son candidatos a COMPETIDOR MODELO: negocios que facturan fuerte con POCOS productos héroe "
        f"(lo contrario a la cola larga tipo TODOMICRO, que no sirve como modelo).\n"
        f"Por cada vendedor: revenue_mes, unidades_mes, publicaciones, top1_share, top3_share y sus 3 "
        f"productos estrella con unidades, revenue y precio_real (monto ÷ unidades).\n\n"
        f"{sellers_str}\n\n"
        f"CONTEXTO DEL ANALISTA:\n"
        f"- Ya importa por courier aéreo desde China (proveedores de fábrica directa, FOB reales "
        f"~1/3 del listado de Alibaba) y vende en MeLi: sensores de CO y protectores de tensión\n"
        f"- Busca productos LIVIANOS (<1kg) para courier; voluminosos/pesados no le sirven\n"
        f"- REGLA COURIER 4×: precio de venta ARS ≥ ~4× el FOB USD en miles. "
        f"Cálculo: neto = precio × 0,85 − $7.000 envío; landed = FOB × 1,975 × TC (${tc:,.0f} ARS/USD)\n"
        f"- Los nicknames vienen ANONIMIZADOS por Nubimetrics: el vendedor real se identifica "
        f"buscando el título del producto estrella en MeLi\n\n"
        f"TAREA: rankeá los mejores 5 a 8 competidores MODELO. Por cada uno:\n"
        f"### [nickname anonimizado]\n"
        f"- **Negocio**: revenue, publicaciones, concentración — qué tipo de jugada es "
        f"(producto héroe + variantes, marca propia, escalera de precios, consumibles)\n"
        f"- **Producto estrella**: qué es, unidades/mes, precio real, ¿traíble por courier? "
        f"(peso estimado, FOB estimado de fábrica, verificación 4×, margen estimado)\n"
        f"- **Riesgos**: certificaciones, marca registrada, exclusividad, precio commodity\n"
        f"- **Veredicto**: 🟢 Modelo a estudiar / 🟡 Evaluar / 🔴 Descartar (motivo)\n\n"
        f"Al final:\n"
        f"1. Tabla resumen: | Vendedor | Revenue | Producto estrella | Unid/mes | Courier | Veredicto |\n"
        f"2. Sección '📋 Próximos pasos': para los 🟢, recomendá identificar al vendedor real en MeLi "
        f"(buscando el título) y exportar su catálogo completo de Nubimetrics para analizarlo con el "
        f"módulo ⚔️ Competidores.\n"
        f"Usá formatos abreviados ($5.9M, 1.2k) y sé concreto."
    )

    def generate():
        try:
            client = anthropic.Anthropic(api_key=api_key)
            chunks = []
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=4000,
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                for text in stream.text_stream:
                    chunks.append(text)
                    yield " "  # keep-alive: evita timeout del proxy de Railway
            yield json.dumps({"analysis": "".join(chunks)})
        except anthropic.APIError as e:
            yield json.dumps({"error": str(e)})

    return Response(
        generate(),
        mimetype="application/json",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


# ---------------------------------------------------------------------------
# mis productos: stock depósito/Full, precio final, ventas 30d, posicionamiento
# ---------------------------------------------------------------------------

def _title_key(title):
    tokens = sorted(w for w in re.split(r"[^\wáéíóúüñ]+", (title or "").lower()) if len(w) >= 3)
    return " ".join(tokens)


def _group_store_items(items):
    """Agrupa publicaciones del mismo producto: comparten inventory_id,
    catalog_product_id o título normalizado (una pub Full + una propia = una fila)."""
    groups = []
    by_key = {}
    for it in items:
        keys = [("tit", _title_key(it["title"]))]
        if it.get("inventory_id"):
            keys.append(("inv", it["inventory_id"]))
        if it.get("catalog_product_id"):
            keys.append(("cat", it["catalog_product_id"]))
        group = next((by_key[k] for k in keys if k in by_key), None)
        if group is None:
            group = {"items": []}
            groups.append(group)
        group["items"].append(it)
        for k in keys:
            by_key[k] = group
    return groups


def _reconcile_product_config(groups):
    """Asocia cada grupo con su fila de product_config por intersección de item_ids.

    Si una fila de config abarca varios grupos heurísticos (merge manual), los une.
    Grupos sin config → INSERT con keyword derivada del título principal.
    """
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = [dict(r) for r in conn.execute("SELECT * FROM product_config ORDER BY id")]
        for r in rows:
            r["_ids"] = set(json.loads(r["item_ids"] or "[]"))

        merged = []
        remaining = list(groups)
        for r in rows:
            mine = [g for g in remaining if r["_ids"] & {i["id"] for i in g["items"]}]
            if not mine:
                continue
            union_group = {"items": [i for g in mine for i in g["items"]], "config": r}
            remaining = [g for g in remaining if g not in mine]
            merged.append(union_group)
            all_ids = sorted(r["_ids"] | {i["id"] for i in union_group["items"]})
            if set(all_ids) != r["_ids"]:
                conn.execute(
                    "UPDATE product_config SET item_ids = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (json.dumps(all_ids), r["id"]),
                )

        for g in remaining:
            main = max(g["items"], key=lambda i: i.get("revenue", 0))
            kw = derive_keyword(main["title"])
            ids = sorted(i["id"] for i in g["items"])
            cur = conn.execute(
                "INSERT INTO product_config (item_ids, keyword) VALUES (?, ?)",
                (json.dumps(ids), kw),
            )
            g["config"] = {
                "id": cur.lastrowid, "keyword": kw, "landed_cost_ars": None, "costo_usd": None,
                "proyeccion_mes": None, "position": None, "position_total": None,
                "position_ts": None, "position_competitors": None,
            }
            merged.append(g)
    return merged


@app.route("/mis-productos")
@login_required
def mis_productos_page():
    return render_template("mis_productos.html")


@app.route("/potencia-ventas")
@login_required
def potencia_ventas_page():
    return render_template("potencia_ventas.html")


@app.route("/api/mis-productos")
@login_required
def mis_productos_data():
    from datetime import datetime, timezone, timedelta

    items, error = get_my_store_items()
    if error:
        return jsonify({"error": error}), 502
    items = [i for i in items if i["status"] == "active"]
    if not items:
        return jsonify({"products": [], "as_of": ""})

    user_id = get_my_user_id()
    tz_arg = timezone(timedelta(hours=-3))
    now = datetime.now(tz_arg)
    fmt = '%Y-%m-%dT%H:%M:%S.000-0300'
    sales = {}
    if user_id:
        sales = fetch_sales_by_item(user_id, (now - timedelta(days=30)).strftime(fmt), now.strftime(fmt))

    prices = get_items_prices([i["id"] for i in items])

    # Stock Full por inventory (deduplicado: varias pubs pueden compartir inventario)
    full_invs = {i["inventory_id"] for i in items if i["logistic_type"] == "fulfillment" and i["inventory_id"]}
    full_stock = {inv: get_fulfillment_stock(inv) for inv in full_invs}

    groups = _reconcile_product_config(_group_store_items(items))

    products = []
    for g in groups:
        cfg = g["config"]
        its = g["items"]
        main = max(its, key=lambda i: i.get("revenue", 0))

        stock_deposito = sum(i["available_quantity"] for i in its if i["logistic_type"] != "fulfillment")
        invs = {i["inventory_id"] for i in its if i["logistic_type"] == "fulfillment" and i["inventory_id"]}
        stock_full = 0
        for inv in invs:
            if full_stock.get(inv) is not None:
                stock_full += full_stock[inv]
            else:  # fallback: available_quantity de la pub Full de ese inventario
                stock_full += max((i["available_quantity"] for i in its if i["inventory_id"] == inv), default=0)
        stock_full += sum(i["available_quantity"] for i in its
                          if i["logistic_type"] == "fulfillment" and not i["inventory_id"])

        price = prices.get(main["id"], {}).get("standard") or main["price"]
        deals = [prices[i["id"]]["deal"] for i in its if prices.get(i["id"], {}).get("deal")]
        sale_price = min(deals) if deals else None

        units_30d = sum(sales.get(i["id"], {}).get("units", 0) for i in its)
        revenue_30d = sum(sales.get(i["id"], {}).get("revenue", 0) for i in its)
        proyeccion = cfg.get("proyeccion_mes")

        products.append({
            "group_id": cfg["id"],
            "title": main["title"],
            "thumbnail": main["thumbnail"],
            "permalink": main["permalink"],
            "items": [{"id": i["id"], "logistic": i["logistic_type"], "stock": i["available_quantity"]} for i in its],
            "stock_deposito": stock_deposito,
            "stock_full": stock_full,
            "price": price,
            "sale_price": sale_price,
            "sales_30d_units": units_30d,
            "sales_30d_revenue": revenue_30d,
            "keyword": cfg.get("keyword") or "",
            "position": cfg.get("position"),
            "position_total": cfg.get("position_total"),
            "position_ts": cfg.get("position_ts"),
            "landed_cost_ars": cfg.get("landed_cost_ars"),
            "proyeccion_mes": proyeccion,
            "needs_strategy": bool(proyeccion) and units_30d < 0.5 * proyeccion,
        })

    products.sort(key=lambda p: -p["sales_30d_revenue"])
    return jsonify({"products": products, "as_of": now.strftime('%H:%M')})


@app.route("/api/mis-productos/config", methods=["POST"])
@login_required
def mis_productos_config():
    data = request.get_json() or {}
    group_id = data.get("group_id")
    if not group_id:
        return jsonify({"error": "Falta group_id"}), 400

    fields = {}
    if "proyeccion_mes" in data:
        v = data["proyeccion_mes"]
        fields["proyeccion_mes"] = int(v) if v not in (None, "") else None
    if "landed_cost_ars" in data:
        v = data["landed_cost_ars"]
        fields["landed_cost_ars"] = float(v) if v not in (None, "") else None
    if "costo_usd" in data:
        v = data["costo_usd"]
        fields["costo_usd"] = float(v) if v not in (None, "") else None
    if "keyword" in data:
        fields["keyword"] = (data["keyword"] or "").strip().lower()
    if not fields:
        return jsonify({"error": "Nada para actualizar"}), 400

    sets = ", ".join(f"{k} = ?" for k in fields)
    vals = list(fields.values())
    if "keyword" in fields:  # keyword nueva invalida el caché de posición
        sets += ", position = NULL, position_total = NULL, position_ts = NULL, position_competitors = NULL"
    with sqlite3.connect(DB_PATH) as conn:
        cur = conn.execute(
            f"UPDATE product_config SET {sets}, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            vals + [group_id],
        )
    if cur.rowcount == 0:
        return jsonify({"error": "Producto no encontrado"}), 404
    return jsonify({"ok": True})


@app.route("/api/mis-productos/merge", methods=["POST"])
@login_required
def mis_productos_merge():
    data = request.get_json() or {}
    keep_id, merge_id = data.get("keep_id"), data.get("merge_id")
    if not keep_id or not merge_id or keep_id == merge_id:
        return jsonify({"error": "IDs inválidos"}), 400
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        keep = conn.execute("SELECT item_ids FROM product_config WHERE id = ?", (keep_id,)).fetchone()
        merge = conn.execute("SELECT item_ids FROM product_config WHERE id = ?", (merge_id,)).fetchone()
        if not keep or not merge:
            return jsonify({"error": "Producto no encontrado"}), 404
        union = sorted(set(json.loads(keep["item_ids"])) | set(json.loads(merge["item_ids"])))
        conn.execute(
            "UPDATE product_config SET item_ids = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (json.dumps(union), keep_id),
        )
        conn.execute("DELETE FROM product_config WHERE id = ?", (merge_id,))
    return jsonify({"ok": True})


@app.route("/api/mis-productos/position", methods=["POST"])
@login_required
def mis_productos_position():
    from datetime import datetime, timezone, timedelta

    data = request.get_json() or {}
    group_id = data.get("group_id")
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM product_config WHERE id = ?", (group_id,)).fetchone()
    if not row:
        return jsonify({"error": "Producto no encontrado"}), 404
    keyword = (row["keyword"] or "").strip()
    if not keyword:
        return jsonify({"error": "Definí una keyword para este producto primero."}), 400

    item_ids = json.loads(row["item_ids"] or "[]")
    seller_id = get_my_user_id()
    catalog_ids = get_items_catalog_ids(item_ids)
    position, total, competitors = get_item_position(item_ids, seller_id, keyword, catalog_ids)
    if total == 0:
        return jsonify({"error": "La búsqueda no devolvió resultados. Probá de nuevo o ajustá la keyword."}), 502

    ts = datetime.now(timezone(timedelta(hours=-3))).strftime('%Y-%m-%dT%H:%M:%S')
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            "UPDATE product_config SET position = ?, position_total = ?, position_ts = ?, "
            "position_competitors = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (position, total, ts, json.dumps(competitors, ensure_ascii=False), group_id),
        )
    return jsonify({"position": position, "total": total, "ts": ts})


@app.route("/api/mis-productos/estrategia", methods=["POST"])
@login_required
def mis_productos_estrategia():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    data = request.get_json() or {}
    product = data.get("product") or {}
    group_id = data.get("group_id") or product.get("group_id")
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM product_config WHERE id = ?", (group_id,)).fetchone()
    if not row:
        return jsonify({"error": "Producto no encontrado"}), 404
    competitors = json.loads(row["position_competitors"] or "[]")
    if not competitors:
        return jsonify({"error": "Actualizá primero el posicionamiento de este producto (botón ↻)."}), 400

    price = float(product.get("price") or 0)
    sale_price = float(product.get("sale_price") or 0) or None
    precio_final = sale_price or price
    landed = row["landed_cost_ars"]
    proyeccion = row["proyeccion_mes"]
    units_30d = int(product.get("sales_30d_units") or 0)
    stock_dep = int(product.get("stock_deposito") or 0)
    stock_full = int(product.get("stock_full") or 0)

    # Margen precalculado acá: la IA no debe inventar aritmética
    envio = 7000 if precio_final >= 33000 else 0
    if landed and precio_final:
        margen = precio_final * 0.85 - envio - landed
        margen_txt = (f"${margen:,.0f} ARS/u ({margen / precio_final * 100:.0f}% del precio) "
                      f"[= {precio_final:,.0f}×0,85 − {envio:,} envío − {landed:,.0f} landed]")
    else:
        margen_txt = "desconocido (falta cargar el costo landed)"

    comp_lines = "\n".join(
        f"{'→ VOS  ' if c.get('is_mine') else '       '}#{c['pos']:>2} | ${c['price']:>12,.0f} | "
        f"{c['title'][:70]} | {c.get('seller_name', '')}"
        + (" | envío gratis" if c.get("free_shipping") else "")
        for c in competitors
    )
    pos_txt = f"#{row['position']} de {row['position_total']}" if row["position"] else \
              f"no aparece entre los primeros {row['position_total']} resultados"

    prompt = (
        "Sos un experto en ventas en MercadoLibre Argentina. Asesorás a un vendedor que "
        "importa por courier desde China (proveedores de fábrica directa) y vende con margen controlado.\n\n"
        "PRODUCTO BAJO ANÁLISIS:\n"
        f"- Título: {product.get('title', '')}\n"
        f"- Precio de lista: ${price:,.0f} ARS"
        + (f" | Precio final con promo activa: ${sale_price:,.0f} ARS\n" if sale_price else " (sin promoción activa)\n")
        + f"- Stock: {stock_dep} u en depósito propio + {stock_full} u en Full\n"
        f"- Ventas últimos 30 días: {units_30d} unidades (${float(product.get('sales_30d_revenue') or 0):,.0f} ARS)\n"
        + (f"- Proyección del vendedor: {proyeccion} u/mes → viene vendiendo al "
           f"{units_30d / proyeccion * 100:.0f}% de lo proyectado\n" if proyeccion else "")
        + (f"- Costo landed (puesto en Argentina): ${landed:,.0f} ARS/u\n" if landed else "")
        + f"- Margen actual por unidad: {margen_txt}\n"
        "- Estructura de costos: comisión MeLi 15%; envío a cargo del vendedor $7.000 si precio ≥ $33.000\n\n"
        f"POSICIONAMIENTO en la búsqueda \"{row['keyword']}\" (medido {row['position_ts'] or 's/d'}):\n"
        f"Posición: {pos_txt}. Primeros resultados reales del listado:\n{comp_lines}\n\n"
        f"TAREA — estrategia concreta para {'llegar a ' + str(proyeccion) + ' u/mes' if proyeccion else 'vender más'}:\n"
        "1. **Diagnóstico**: ¿el problema es precio, posición, competencia, título o logística?\n"
        "2. **Precio**: compará contra los competidores del listado real. Si sugerís descuento o "
        "nuevo precio, calculá el margen resultante en $/u y % usando la estructura de costos de arriba. "
        "NUNCA sugieras un precio que deje margen menor al 20% del precio de venta"
        + (" (el landed es $%s)" % f"{landed:,.0f}" if landed else "") + ".\n"
        "3. **Posicionamiento**: mejoras concretas de título/keyword, si conviene Product Ads "
        "(estimá el ACOS tolerable con el margen disponible), y si conviene mandar stock a Full "
        "(impacto en ranking y buy box).\n"
        "4. **Plan de acción priorizado**: 3 a 5 pasos concretos, cada uno con impacto esperado.\n\n"
        "Formato: markdown con ### por sección, veredicto 🟢🟡🔴 en cada palanca. "
        "Montos abreviados ($85k, $1,2M). Sé directo y accionable, sin relleno."
    )

    def generate():
        try:
            client = anthropic.Anthropic(api_key=api_key)
            chunks = []
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=3000,
                messages=[{"role": "user", "content": prompt}],
            ) as stream:
                for text in stream.text_stream:
                    chunks.append(text)
                    yield " "  # keep-alive: evita timeout del proxy de Railway
            yield json.dumps({"analysis": "".join(chunks)})
        except anthropic.APIError as e:
            yield json.dumps({"error": str(e)})

    return Response(
        generate(),
        mimetype="application/json",
        headers={"X-Accel-Buffering": "no", "Cache-Control": "no-cache"},
    )


# ---------------------------------------------------------------------------
# Potencia tus Ventas: stock + ventas del período + meta de facturación → estrategia IA
# ---------------------------------------------------------------------------

@app.route("/api/potencia-datos")
@login_required
def potencia_datos():
    from datetime import datetime, timezone, timedelta
    import calendar

    period = request.args.get("period", "30d")
    tc = float(request.args.get("tc") or 0)
    tz_arg = timezone(timedelta(hours=-3))
    now = datetime.now(tz_arg)
    fmt = '%Y-%m-%dT%H:%M:%S.000-0300'

    items, error = get_my_store_items()
    if error:
        return jsonify({"error": error}), 502
    items = [i for i in items if i["status"] == "active"]

    user_id = get_my_user_id()
    if not user_id:
        return jsonify({"error": "No se pudo obtener el usuario de MeLi."}), 502

    month_from = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    now_str = now.strftime(fmt)
    facturado_mes, _ = fetch_orders_total(user_id, month_from.strftime(fmt), now_str)
    last_day = calendar.monthrange(now.year, now.month)[1]
    dias_restantes = last_day - now.day + 1

    if not items:
        return jsonify({
            "products": [], "period": period, "facturado_mes": facturado_mes,
            "dias_restantes": dias_restantes, "as_of": now.strftime('%H:%M'),
        })

    date_from = month_from if period == "month" else now - timedelta(days=30)
    sales = fetch_sales_by_item(user_id, date_from.strftime(fmt), now_str)
    prices = get_items_prices([i["id"] for i in items])

    full_invs = {i["inventory_id"] for i in items if i["logistic_type"] == "fulfillment" and i["inventory_id"]}
    full_stock = {inv: get_fulfillment_stock(inv) for inv in full_invs}

    groups = _reconcile_product_config(_group_store_items(items))

    products = []
    for g in groups:
        its = g["items"]
        main = max(its, key=lambda i: i.get("revenue", 0))

        stock_deposito = sum(i["available_quantity"] for i in its if i["logistic_type"] != "fulfillment")
        invs = {i["inventory_id"] for i in its if i["logistic_type"] == "fulfillment" and i["inventory_id"]}
        stock_full = 0
        for inv in invs:
            if full_stock.get(inv) is not None:
                stock_full += full_stock[inv]
            else:  # fallback: available_quantity de la pub Full de ese inventario
                stock_full += max((i["available_quantity"] for i in its if i["inventory_id"] == inv), default=0)
        stock_full += sum(i["available_quantity"] for i in its
                          if i["logistic_type"] == "fulfillment" and not i["inventory_id"])

        price = prices.get(main["id"], {}).get("standard") or main["price"]
        deals = [prices[i["id"]]["deal"] for i in its if prices.get(i["id"], {}).get("deal")]
        sale_price = min(deals) if deals else None

        units_periodo = sum(sales.get(i["id"], {}).get("units", 0) for i in its)
        revenue_periodo = sum(sales.get(i["id"], {}).get("revenue", 0) for i in its)

        iva_pct = main.get("iva_pct")
        costo_usd = g["config"].get("costo_usd")
        landed_cost_ars = None
        if costo_usd and tc:
            landed_cost_ars = costo_usd * tc * (1 + (iva_pct or 0) / 100)

        products.append({
            "group_id": g["config"]["id"],
            "title": main["title"],
            "thumbnail": main["thumbnail"],
            "permalink": main["permalink"],
            "stock_deposito": stock_deposito,
            "stock_full": stock_full,
            "price": price,
            "sale_price": sale_price,
            "units_periodo": units_periodo,
            "revenue_periodo": revenue_periodo,
            "iva_pct": iva_pct,
            "costo_usd": costo_usd,
            "landed_cost_ars": landed_cost_ars,
        })

    products.sort(key=lambda p: -p["revenue_periodo"])

    return jsonify({
        "products": products,
        "period": period,
        "facturado_mes": facturado_mes,
        "dias_restantes": dias_restantes,
        "as_of": now.strftime('%H:%M'),
    })


@app.route("/api/potencia-analyze", methods=["POST"])
@login_required
def potencia_analyze():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada."}), 500

    data = request.get_json() or {}
    products = data.get("products", [])
    target_revenue = data.get("target_revenue", 0)
    facturado_mes = data.get("facturado_mes", 0)
    dias_restantes = data.get("dias_restantes", 1) or 1
    period_label = "mes actual" if data.get("period") == "month" else "últimos 30 días"

    if not products:
        return jsonify({"error": "Sin datos de stock/ventas para analizar."}), 400

    gap = target_revenue - facturado_mes
    top_products = sorted(products, key=lambda p: -p.get("revenue_periodo", 0))[:60]

    def margen_info(p):
        precio_final = p.get("sale_price") or p["price"]
        landed = p.get("landed_cost_ars")
        if not landed or not precio_final:
            return "sin costo cargado"
        envio = 7000 if precio_final >= 33000 else 0
        margen = precio_final * 0.85 - envio - landed
        return f"${margen:,.0f} ARS/u ({margen / precio_final * 100:.0f}% del precio de venta)"

    catalog = [
        {
            "titulo": p["title"],
            "precio_lista_ARS": p["price"],
            "precio_con_promo_ARS": p.get("sale_price"),
            "stock_deposito_propio": p["stock_deposito"],
            "stock_full": p["stock_full"],
            "unidades_vendidas_periodo": p["units_periodo"],
            "revenue_periodo_ARS": p["revenue_periodo"],
            "costo_landed_ARS_por_unidad": p.get("landed_cost_ars"),
            "margen_actual_precalculado": margen_info(p),
        }
        for p in top_products
    ]

    gap_line = (
        f"Falta facturar ${gap:,.0f} ARS (${gap / dias_restantes:,.0f} ARS/día necesarios)."
        if gap > 0
        else f"Ya se superó el objetivo por ${abs(gap):,.0f} ARS."
    )

    prompt = (
        "Sos un experto en e-commerce y gestión de inventario en MercadoLibre Argentina.\n"
        f"Te paso el catálogo activo de un vendedor: stock real separado entre depósito propio y el "
        f"centro Full de MercadoLibre, precio (de lista y con promo activa si tiene), ventas de los "
        f"{period_label} unidad por unidad, y el margen actual ya calculado cuando el vendedor cargó "
        f"el costo del producto (costo en USD sin IVA → convertido a ARS con el tipo de cambio e IVA "
        f"real de cada publicación).\n\n"
        f"Facturación acumulada este mes: ${facturado_mes:,.0f} ARS.\n"
        f"Objetivo total de facturación para lo que resta del mes: ${target_revenue:,.0f} ARS.\n"
        f"Faltan {dias_restantes} días para cerrar el mes.\n"
        f"{gap_line}\n\n"
        "IMPORTANTE sobre el margen: 'margen_actual_precalculado' ya viene calculado (comisión MeLi 15%, "
        "envío $7.000 a cargo del vendedor si el precio ≥ $33.000, menos el costo landed real). NO "
        "recalcules ni inventes esta aritmética — usalo tal cual viene. Para productos con "
        "'sin costo cargado', no tenés margen real: no asumas uno, y si vas a sugerir descuento en esos "
        "casos aclará que falta cargar el costo para confirmar que el precio con descuento sigue siendo rentable.\n\n"
        "Con esta información armá una **estrategia concreta y priorizada**:\n"
        "1. **Productos para bajar de precio / hacer descuento**: stock total (depósito + Full) alto "
        "y rotación baja en el período. Sugerí % de descuento aproximado y el motivo, verificando con el "
        "margen precalculado que el descuento no deje el margen negativo o por debajo del 15-20%.\n"
        "2. **Productos para enviar a Full**: tienen stock en depósito propio y buena rotación en el "
        "período — priorizalos porque Full mejora visibilidad, buy box y tiempos de envío. Si ya tienen "
        "algo de stock en Full pero se está por agotar, marcalo también acá.\n"
        "3. **Productos a mantener/reforzar**: buena rotación y ya bien abastecidos en Full — mantener "
        "y eventualmente reponer.\n"
        "4. **Riesgo de quiebre de stock**: alta rotación con pocas unidades restantes (depósito + Full) "
        "para lo que queda del mes.\n\n"
        "Cerrá con un **Resumen ejecutivo**: 3-5 acciones priorizadas para esta semana y una conclusión "
        "de si, aplicando la estrategia, se puede alcanzar el objetivo de facturación restante.\n\n"
        f"Catálogo (top {len(catalog)} productos por revenue del período):\n"
        f"{json.dumps(catalog, ensure_ascii=False, indent=1)}"
    )

    try:
        client = anthropic.Anthropic(api_key=api_key)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=4000,
            messages=[{"role": "user", "content": prompt}],
        )
        return jsonify({"analysis": response.content[0].text})
    except anthropic.APIError as e:
        return jsonify({"error": str(e)}), 502


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
